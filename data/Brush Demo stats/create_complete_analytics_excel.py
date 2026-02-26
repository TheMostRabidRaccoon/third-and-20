#!/usr/bin/env python3
"""Create comprehensive Excel workbook for Brush complete analytics"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from pathlib import Path

BASE = Path("/sessions/exciting-friendly-curie/mnt/rri_server/data/Brush Demo stats")
CV_OUTPUT = Path("/sessions/exciting-friendly-curie/mnt/rri_server/projects/third_and_20/output")

# Styles
HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SECTION_FILL = PatternFill(fill_type="solid", fgColor="D6DCE5")
SECTION_FONT = Font(bold=True, size=11)
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def style_header_row(sheet, row, num_cols):
    for col in range(1, num_cols + 1):
        cell = sheet.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = THIN_BORDER

def auto_width(sheet, df):
    for i, col in enumerate(df.columns, 1):
        max_len = max(len(str(col)), df[col].astype(str).str.len().max() if len(df) > 0 else 0)
        sheet.column_dimensions[get_column_letter(i)].width = min(max_len + 2, 20)

def create_workbook():
    wb = Workbook()

    # Sheet 1: Game Analytics Summary
    ws1 = wb.active
    ws1.title = "Game Analytics"
    df1 = pd.read_csv(BASE / "brush_COMPLETE_game_analytics.csv")

    # Select key columns for summary view
    summary_cols = ['opponent', 'total_plays', 'total_yards', 'total_tds', 'yards_per_play',
                    'first_downs', 'completion_pct', 'rushing_yards', 'passing_yards',
                    'def_total_yards', 'def_yards_per_play', 'third_down_pct',
                    'cv_run_pct', 'cv_top_formation', 'cv_tackles', 'cv_tackle_efficiency',
                    'avg_sdi']
    summary_cols = [c for c in summary_cols if c in df1.columns]
    df_summary = df1[summary_cols]

    for r_idx, row in enumerate(dataframe_to_rows(df_summary, index=False, header=True), 1):
        for c_idx, val in enumerate(row, 1):
            cell = ws1.cell(row=r_idx, column=c_idx, value=val)
            cell.border = THIN_BORDER
            if r_idx == 1:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = Alignment(horizontal='center', wrap_text=True)

    # Add totals row
    total_row = len(df_summary) + 2
    ws1.cell(row=total_row, column=1, value="SEASON TOTALS").font = SECTION_FONT
    ws1.cell(row=total_row, column=2, value=f'=SUM(B2:B{total_row-1})')
    ws1.cell(row=total_row, column=3, value=f'=SUM(C2:C{total_row-1})')
    ws1.cell(row=total_row, column=4, value=f'=SUM(D2:D{total_row-1})')
    ws1.cell(row=total_row, column=5, value=f'=AVERAGE(E2:E{total_row-1})')

    auto_width(ws1, df_summary)

    # Sheet 2: Full Traditional Stats
    ws2 = wb.create_sheet("Traditional Stats")
    trad_cols = [c for c in df1.columns if not c.startswith('cv_') and c not in ['avg_sdi', 'sdi_std', 'max_sdi', 'min_sdi', 'sdi_plays']]
    df_trad = df1[trad_cols]

    for r_idx, row in enumerate(dataframe_to_rows(df_trad, index=False, header=True), 1):
        for c_idx, val in enumerate(row, 1):
            cell = ws2.cell(row=r_idx, column=c_idx, value=val)
            cell.border = THIN_BORDER
            if r_idx == 1:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT

    # Sheet 3: CV Cognitive Metrics
    ws3 = wb.create_sheet("CV Cognitive Metrics")
    cv_cols = ['opponent'] + [c for c in df1.columns if c.startswith('cv_') or c in ['avg_sdi', 'sdi_std', 'max_sdi', 'min_sdi', 'sdi_plays']]
    df_cv = df1[cv_cols]

    for r_idx, row in enumerate(dataframe_to_rows(df_cv, index=False, header=True), 1):
        for c_idx, val in enumerate(row, 1):
            cell = ws3.cell(row=r_idx, column=c_idx, value=val)
            cell.border = THIN_BORDER
            if r_idx == 1:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT

    auto_width(ws3, df_cv)

    # Sheet 4: Defensive Individual Stats
    ws4 = wb.create_sheet("Defense Individual")
    if (BASE / "brush_COMPLETE_defense_individual.csv").exists():
        df_def = pd.read_csv(BASE / "brush_COMPLETE_defense_individual.csv")
        for r_idx, row in enumerate(dataframe_to_rows(df_def, index=False, header=True), 1):
            for c_idx, val in enumerate(row, 1):
                cell = ws4.cell(row=r_idx, column=c_idx, value=val)
                cell.border = THIN_BORDER
                if r_idx == 1:
                    cell.fill = HEADER_FILL
                    cell.font = HEADER_FONT
        auto_width(ws4, df_def)

    # Sheet 5: SDI Player Rankings
    ws5 = wb.create_sheet("SDI Rankings")
    if (CV_OUTPUT / "01_SDI_Player_Rankings.csv").exists():
        df_sdi = pd.read_csv(CV_OUTPUT / "01_SDI_Player_Rankings.csv")
        for r_idx, row in enumerate(dataframe_to_rows(df_sdi, index=False, header=True), 1):
            for c_idx, val in enumerate(row, 1):
                cell = ws5.cell(row=r_idx, column=c_idx, value=val)
                cell.border = THIN_BORDER
                if r_idx == 1:
                    cell.fill = HEADER_FILL
                    cell.font = HEADER_FONT
        auto_width(ws5, df_sdi)

    # Sheet 6: Offensive Tendencies
    ws6 = wb.create_sheet("Offensive Tendencies")
    if (CV_OUTPUT / "02_Offensive_Tendencies_By_Opponent.csv").exists():
        df_tend = pd.read_csv(CV_OUTPUT / "02_Offensive_Tendencies_By_Opponent.csv")
        for r_idx, row in enumerate(dataframe_to_rows(df_tend, index=False, header=True), 1):
            for c_idx, val in enumerate(row, 1):
                cell = ws6.cell(row=r_idx, column=c_idx, value=val)
                cell.border = THIN_BORDER
                if r_idx == 1:
                    cell.fill = HEADER_FILL
                    cell.font = HEADER_FONT
        auto_width(ws6, df_tend)

    # Sheet 7: Formation Analysis
    ws7 = wb.create_sheet("Formations")
    if (BASE / "brush_hudl_formation_tendencies.csv").exists():
        df_form = pd.read_csv(BASE / "brush_hudl_formation_tendencies.csv")
        for r_idx, row in enumerate(dataframe_to_rows(df_form, index=False, header=True), 1):
            for c_idx, val in enumerate(row, 1):
                cell = ws7.cell(row=r_idx, column=c_idx, value=val)
                cell.border = THIN_BORDER
                if r_idx == 1:
                    cell.fill = HEADER_FILL
                    cell.font = HEADER_FONT
        auto_width(ws7, df_form)

    # Save
    output_path = BASE / "Brush_COMPLETE_Analytics_Workbook.xlsx"
    wb.save(output_path)
    print(f"Created: {output_path}")
    return output_path

if __name__ == "__main__":
    create_workbook()
